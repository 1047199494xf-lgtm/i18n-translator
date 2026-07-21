#!/usr/bin/env python3
"""i18n 翻译工具 — 本地服务后端 v2"""
import re, json, time, os, io
from pathlib import Path
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import openpyxl
import requests

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

dict_zh_to_all = {}     # zh → {lang: translation}
dict_all = {}           # {lang: {zh: translation}}  每个语言独立的查找表
dict_total_rows = 0
dict_lang_cols = {}     # 字典里有哪些语言列: {lang_code: col_index}
api_key = ''
api_base = 'https://api.deepseek.com'

# ═══ API Key 持久化 ═══
CONFIG_FILE = os.path.join(os.path.dirname(__file__), 'config.json')

def load_config():
    global api_key, api_base
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
                api_key = cfg.get('key', '')
                api_base = cfg.get('base', 'https://api.deepseek.com')
        except: pass

def save_config():
    with open(CONFIG_FILE, 'w') as f:
        json.dump({'key': api_key, 'base': api_base}, f)

load_config()  # 启动时加载

LANG_NAMES = {
    'en':'English','fr':'Français','ar':'العربية','mn':'Монгол',
    'ja':'日本語','ko':'한국어','ru':'Русский','es':'Español',
    'de':'Deutsch','pt':'Português','it':'Italiano','th':'ไทย',
    'vi':'Tiếng Việt','tr':'Türkçe','nl':'Nederlands','pl':'Polski',
    'uk':'Українська','he':'עברית','id':'Bahasa Indonesia','ms':'Melayu',
    'hi':'हिन्दी','bn':'বাংলা','fil':'Filipino','sw':'Kiswahili',
    'km':'ខ្មែរ','lo':'ລາວ','my':'မြန်မာ',
}

# ═══════════════════════════════════════
# 字典
# ═══════════════════════════════════════

def load_dict_xlsx(filepath):
    global dict_zh_to_all, dict_all, dict_total_rows, dict_lang_cols
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return 0

    header = [str(c).strip() if c else '' for c in rows[0]]

    # 识别所有语言列
    LANG_KEYS = {'zh','中文','chinese','en','english','英文','fr','french','法文',
                 'ar','arabic','阿拉伯','mn','mongolian','蒙古','ja','japanese','日文',
                 'ko','korean','韩文','ru','russian','俄文','es','spanish','西文',
                 'de','german','德文','pt','portuguese','it','italian','th','thai',
                 'vi','vietnamese','tr','turkish'}
    LANG_MAP = {'中文':'zh','chinese':'zh','english':'en','英文':'en','french':'fr',
                '法文':'fr','arabic':'ar','阿拉伯':'ar','mongolian':'mn','蒙古':'mn',
                'japanese':'ja','日文':'ja','korean':'ko','韩文':'ko','russian':'ru',
                '俄文':'ru','spanish':'es','西文':'es','german':'de','德文':'de',
                'portuguese':'pt','italian':'it','thai':'th','vietnamese':'vi',
                'turkish':'tr'}

    zh_col = None
    dict_lang_cols = {}
    for i, h in enumerate(header):
        hl = h.lower().strip()
        if hl in ('zh','中文','chinese'): zh_col = i
        elif hl in LANG_MAP:
            code = LANG_MAP.get(hl, hl)
            dict_lang_cols[code] = i
        elif hl in LANG_KEYS:
            dict_lang_cols[hl] = i

    # 也尝试从 header 直接识别两字母代码
    for i, h in enumerate(header):
        hl = h.lower().strip()
        if len(hl) == 2 and hl not in dict_lang_cols and i != zh_col:
            # 可能是语言代码
            if hl in ('en','fr','ar','mn','ja','ko','ru','es','de','pt','it','th','vi','tr','nl','pl','uk','he','id','ms','hi','bn','fil','sw','km','lo','my'):
                dict_lang_cols[hl] = i

    if zh_col is None: return 0

    dict_zh_to_all.clear(); dict_all.clear()
    total = 0
    for row in rows[1:]:
        zh = str(row[zh_col]).strip() if zh_col < len(row) and row[zh_col] else ''
        if not zh: continue

        translations = {}
        for lang, col in dict_lang_cols.items():
            val = str(row[col]).strip() if col < len(row) and row[col] else ''
            if val:
                translations[lang] = val
                if lang not in dict_all:
                    dict_all[lang] = {}
                dict_all[lang][zh] = val

        if translations:
            if zh not in dict_zh_to_all:
                dict_zh_to_all[zh] = {}
            dict_zh_to_all[zh].update(translations)
            total += 1

    dict_total_rows = total
    print(f"字典: {total} 条中文, 语言列: {list(dict_lang_cols.keys())}")
    return dict_total_rows

def get_dict_trans(zh, lang):
    """从字典获取指定语言的翻译"""
    entry = dict_zh_to_all.get(zh, {})
    return entry.get(lang, '') or entry.get('en', '')

def cap_first(t):
    if not t: return t
    return t[0].upper() + t[1:]

def split_translate(text, target_lang='en'):
    """拆词翻译：贪心最长匹配，目标语言优先"""
    if not text: return None
    lang_dict = dict_all.get(target_lang, {})
    en_dict = dict_all.get('en', {})  # fallback

    result = []; i = 0
    while i < len(text):
        best_val = None; best_len = 0
        for j in range(min(i + 15, len(text)), i, -1):
            chunk = text[i:j]
            if chunk in lang_dict:
                best_val = lang_dict[chunk]; best_len = len(chunk)
                break
            elif chunk in en_dict:
                best_val = en_dict[chunk]; best_len = len(chunk)
                break
        if best_val: result.append(best_val); i += best_len
        else: result.append(text[i]); i += 1

    combined = ' '.join(result)
    if re.search(r'[一-鿿]', combined): return None
    words = combined.split()
    singles = [w for w in words if len(w) == 1 and w.isalpha()]
    if len(singles) >= 2: return None
    return combined

def ai_review_dict(zh, lang, trans):
    """AI 快速审查字典翻译是否明显错误，返回修正建议或 None"""
    if not api_key: return None
    lang_name = LANG_NAMES.get(lang, lang)

    prompt = f"""Review this dictionary entry for a gas station management system.

Chinese: "{zh}"
Dictionary translation ({lang_name}): "{trans}"

Check for these common errors:
1. Extra words not in the original (e.g. "起始" -> "Start Time" instead of just "Start")
2. Opposite or wrong meaning (e.g. "增加" -> "Decrease", "停用" -> "Active")
3. Wrong industry term (e.g. "卸油" -> "Uninstall Oil" instead of "Delivery Management")
4. Pinyin output instead of actual translation

Reply ONLY: 'OK' if accurate, or 'FIX: <corrected translation>' if wrong. Optional brief reason."""

    try:
        resp = requests.post(
            f'{api_base}/v1/chat/completions',
            json={
                'model': 'deepseek-chat',
                'messages': [{'role': 'user', 'content': prompt}],
                'temperature': 0.1, 'max_tokens': 100,
            },
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            timeout=15,
        )
        result = resp.json()['choices'][0]['message']['content'].strip()
        if result.upper().startswith('OK'):
            return None  # 没问题
        # 提取修正建议
        fix_match = re.search(r'FIX:\s*(.+?)(?:\.|$)', result)
        if fix_match:
            return {'warning': True, 'dict_value': trans, 'suggestion': fix_match.group(1).strip(), 'reason': result}
        return None
    except:
        return None

def translate_multi(text, target_langs, review=False):
    """多语言翻译: {lang: {value, warning?, dict_value?, suggestion?}}
    review=True 时启用 AI 审查（仅快速翻译模式使用）"""
    results = {}
    for lang in target_langs:
        if lang == 'zh': continue
        result_obj = {'value': text}

        # 1. 精确匹配
        entry = dict_zh_to_all.get(text, {})
        if lang in entry and entry[lang]:
            trans = entry[lang]
            result_obj = {'value': cap_first(trans), 'source': 'dict'}

            # AI 审查（仅在快速翻译模式）
            if review:
                review_result = ai_review_dict(text, lang, trans)
                if review_result:
                    result_obj['warning'] = True
                    result_obj['dict_value'] = trans
                    result_obj['suggestion'] = review_result['suggestion']
                    result_obj['reason'] = review_result['reason']

            results[lang] = result_obj
            continue

        # 2. 英文 fallback
        if 'en' in entry and entry['en']:
            result_obj = {'value': cap_first(entry['en']), 'source': 'dict_en'}
            results[lang] = result_obj
            continue

        # 3. 拆词
        split_result = split_translate(text, lang)
        if split_result:
            result_obj = {'value': cap_first(split_result), 'source': 'split'}
            results[lang] = result_obj
            continue

        # 4. AI 翻译
        if api_key:
            ai_result = ai_translate(text, lang)
            if ai_result:
                result_obj = {'value': cap_first(ai_result), 'source': 'ai'}
                results[lang] = result_obj
                continue

        results[lang] = result_obj

    return results

def ai_translate(text, target_lang):
    """AI 单语言翻译"""
    if not api_key: return None
    lang_name = LANG_NAMES.get(target_lang, target_lang)
    sys_prompt = f"""You are a professional translator specializing in gas station management systems (加油站后台管理系统).

Context: This is a fuel retail management platform covering:
- Fuel product management (油品管理): oil types, pricing, inventory, delivery
- Fleet management (车队管理): fleet accounts, prepayment, credit sales
- Member management (会员管理): loyalty cards, fuel cards, member tiers
- Points management (积分管理): points earning, deduction, redemption
- Payment & settlement (支付/结算): multiple payment methods, shift settlement
- Station equipment (站级设置): pumps, nozzles, tanks, ATG, POS
- Reports & analytics (报表分析): daily/shift reports, consumption statistics
- System maintenance (系统维护): monitoring, alarms, data management

Translation style:
- Be concise and professional, matching the tone of enterprise software
- Use industry-standard abbreviations where appropriate (POS, ATG, IC Card, ETC)
- "加油" → Refueling, "油品" → Oil/Product, "车队" → Fleet, "备用金" → Reserve Fund
- "充值" → Recharge/Top-up, "扣减" → Deduction, "日结" → Daily Closing
- Preserve technical terms: PSAM, NFC, E-Account, CyberView, Fuel-In Card
- French: use formal business French, Arabic: use standard Modern Standard Arabic, Mongolian: use Cyrillic script

Translate the following text to {lang_name}. Output only the translation, no explanation, no markdown."""

    try:
        resp = requests.post(
            f'{api_base}/v1/chat/completions',
            json={
                'model': 'deepseek-chat',
                'messages': [
                    {'role': 'system', 'content': sys_prompt},
                    {'role': 'user', 'content': text},
                ],
                'temperature': 0.3, 'max_tokens': 300,
            },
            headers={'Authorization': f'Bearer {api_key}', 'Content-Type': 'application/json'},
            timeout=30,
        )
        result = resp.json()['choices'][0]['message']['content'].strip()
        return result if not any('一' <= c <= '鿿' for c in result) else None
    except:
        return None

# ═══════════════════════════════════════
# 文件处理
# ═══════════════════════════════════════

def process_js_file(content, target_langs):
    """处理 JS/JSON 文件 —— 批量 AI 翻译版"""
    lines = content.splitlines()
    lang_codes = frozenset({
        'zh','en','mn','ar','fr','ja','ko','ru','es','de',
        'pt','it','th','vi','tr','nl','pl','uk','he','id',
        'ms','hi','bn','fil','sw','km','lo','my',
    })
    zh_re = re.compile(r"""\bzh\s*:\s*(['"])(.*?)\1""", re.IGNORECASE)
    obj_open = re.compile(r"""^\s*(['"]?)([\w]+)\1\s*:\s*\{""")
    lang_re = re.compile(r"""^\s*(['"]?)([\w]{2,5})\1\s*:\s*(['"])(.*?)\3""")

    # 第一遍：收集所有 zh 值
    all_zh = set()
    indent_stack = []
    for i, line in enumerate(lines):
        stripped = line.strip()
        indent = len(line) - len(line.lstrip())
        if stripped in ('}', '},'):
            while indent_stack and indent_stack[-1][0] >= indent:
                indent_stack.pop()
            continue
        zh_m = zh_re.search(line)
        if zh_m:
            all_zh.add(zh_m.group(2))
        obj_m = obj_open.match(line)
        if obj_m:
            key_name = obj_m.group(2)
            while indent_stack and indent_stack[-1][0] >= indent:
                indent_stack.pop()
            indent_stack.append((indent, key_name))

    # 批量翻译所有 zh 值（优先字典，缺失的批量 AI）
    translations = {}  # {zh: {lang: value}}
    untranslated = set()
    for zh_val in all_zh:
        entry = dict_zh_to_all.get(zh_val, {})
        if entry:
            translations[zh_val] = {}
            for lang in target_langs:
                if lang in entry:
                    translations[zh_val][lang] = cap_first(entry[lang])
                elif 'en' in entry:
                    translations[zh_val][lang] = cap_first(entry['en'])
                else:
                    untranslated.add(zh_val)
                    break  # 这个 zh 没完整翻译
        else:
            untranslated.add(zh_val)

    # AI 批量翻译缺失词
    if untranslated and api_key:
        batch_result = ai_batch_translate(list(untranslated), target_langs)
        for zh_val, lang_trans in batch_result.items():
            if zh_val not in translations:
                translations[zh_val] = {}
            for lang, val in lang_trans.items():
                if val:
                    translations[zh_val][lang] = cap_first(val)
        # 补上没 AI 结果的
        for zh_val in untranslated:
            if zh_val not in translations:
                translations[zh_val] = {}

    # 第二遍：应用翻译
    indent_stack = []
    changes = []
    new_lines = []
    seen_changes = set()  # 去重

    for i, line in enumerate(lines):
        stripped = line.strip()
        indent = len(line) - len(line.lstrip())

        if stripped in ('}', '},'):
            while indent_stack and indent_stack[-1][0] >= indent:
                indent_stack.pop()
            continue

        zh_m = zh_re.search(line)
        if zh_m:
            zh_val = zh_m.group(2)
            zh_trans = translations.get(zh_val, {})
            key_parts = [n for _, n in indent_stack if n.lower() not in lang_codes]
            full_key = '.'.join(key_parts) if key_parts else ''

            existing_langs = set()
            j = i + 1
            last_lang_j = i
            while j < len(lines):
                l2 = lines[j]; s2 = l2.strip()
                if not s2 or s2.startswith('//'): j += 1; continue
                l_indent = len(l2) - len(l2.lstrip())
                if s2 in ('}', '},') and l_indent <= indent: break
                lm = lang_re.match(l2)
                if lm and l_indent == indent:
                    code = lm.group(2).lower()
                    existing_langs.add(code)
                    if code in zh_trans and code != 'zh':
                        old_val = lm.group(4)
                        new_val = zh_trans[code]
                        if new_val and new_val != old_val:
                            ck = f'{full_key}.{code}'
                            if ck not in seen_changes:
                                seen_changes.add(ck)
                                changes.append({'key':ck,'old':old_val,'new':new_val,'annotation':zh_val})
                            quote = lm.group(3)
                            pattern = rf"""(\b{re.escape(code)}\s*:\s*{re.escape(quote)})(.*?)({re.escape(quote)})"""
                            l2 = re.sub(pattern, rf'\1{new_val}\3', l2, count=1)
                            lines[j] = l2
                    last_lang_j = j
                j += 1

            missing_langs = [l for l in target_langs if l not in existing_langs and l != 'zh']
            if missing_langs:
                indent_str = line[:len(line)-len(line.lstrip())] + '  '
                for ml in missing_langs:
                    val = zh_trans.get(ml, '')
                    new_line = f"{indent_str}{ml}: '{val}',\n"
                    inserts_after = last_lang_j if 0 <= last_lang_j < len(lines) else i
                    new_lines.append((inserts_after, new_line))
                    ck = f'{full_key}.{ml}'
                    if ck not in seen_changes:
                        seen_changes.add(ck)
                        changes.append({'key':ck,'old':'','new':val,'annotation':zh_val})

        obj_m = obj_open.match(line)
        if obj_m:
            key_name = obj_m.group(2)
            while indent_stack and indent_stack[-1][0] >= indent:
                indent_stack.pop()
            indent_stack.append((indent, key_name))

    result_lines = lines[:]
    for pos, new_line in sorted(new_lines, key=lambda x: x[0], reverse=True):
        result_lines.insert(pos + 1, new_line)

    return '\n'.join(result_lines), changes

def process_sql_file(content, target_langs):
    """处理 SQL 文件"""
    changes = []
    pattern = re.compile(
        r"\('(\d+)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)',\s*'([^']*)'\)",
        re.DOTALL
    )

    result = []; pos = 0
    for m in pattern.finditer(content):
        result.append(content[pos:m.start()])
        fields = list(m.groups())
        item_name = fields[4]
        annotation = fields[11].strip() if len(fields) > 11 else ''

        if item_name and annotation:
            trans = translate_multi(item_name, target_langs)
            en_obj = trans.get('en', {})
            en_trans = en_obj.get('value', '') if isinstance(en_obj, dict) else str(en_obj)
            if en_trans and en_trans.lower() != item_name.lower():
                changes.append({'old': item_name, 'new': en_trans, 'annotation': annotation})
                fields[4] = en_trans

        result.append("('" + "', '".join(fields) + "')")
        pos = m.end()
    result.append(content[pos:])
    return ''.join(result), changes

# ═══════════════════════════════════════
# API
# ═══════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/health')
def health():
    return jsonify({'status':'ok'})

@app.route('/api/set-key', methods=['POST'])
def set_key():
    global api_key, api_base
    data = request.json
    api_key = data.get('key','')
    api_base = data.get('base','https://api.deepseek.com')
    save_config()
    return jsonify({'ok':True,'has_key':bool(api_key)})

@app.route('/api/load-dict', methods=['POST'])
def load_dict():
    if 'file' not in request.files:
        return jsonify({'error':'No file'}), 400
    f = request.files['file']
    path = os.path.join('/tmp', f.filename)
    f.save(path)
    count = load_dict_xlsx(path)
    return jsonify({'ok':True,'count':count,'unique_zh':len(dict_zh_to_all),'langs':list(dict_lang_cols.keys())})

@app.route('/api/get-key')
def get_key():
    return jsonify({'key': api_key[:8]+'...' if len(api_key)>8 else '', 'base': api_base, 'has_key': bool(api_key)})

@app.route('/api/translate-text', methods=['POST'])
def translate_single():
    data = request.json
    text = data.get('text','')
    langs = data.get('langs',['en'])
    results = translate_multi(text, langs, review=True)
    return jsonify({'original':text,'results':results})

@app.route('/api/translate-file', methods=['POST'])
def translate_file():
    if 'file' not in request.files:
        return jsonify({'error':'No file'}), 400

    f = request.files['file']
    content = f.read().decode('utf-8')
    fname = f.filename.lower()

    # 读取目标语言（多选）
    langs_str = request.form.get('langs', 'en')
    target_langs = [l.strip() for l in langs_str.split(',') if l.strip()]

    if fname.endswith(('.js','.json','.txt')):
        new_content, changes = process_js_file(content, target_langs)
        file_type = 'js'
    elif fname.endswith('.sql'):
        new_content, changes = process_sql_file(content, target_langs)
        file_type = 'sql'
    elif fname.endswith(('.xlsx','.xls')):
        # Excel 走专门的翻译端点
        new_content, changes = process_js_file(content, target_langs)
        file_type = 'xlsx'
    else:
        new_content, changes = process_js_file(content, target_langs)
        file_type = 'js'

    return jsonify({
        'ok':True,'type':file_type,
        'changes':len(changes),'details':changes[:100],
        'content':new_content,
    })

@app.route('/api/translate-excel', methods=['POST'])
def translate_excel():
    if 'file' not in request.files:
        return jsonify({'error':'No file'}), 400

    f = request.files['file']
    langs_str = request.form.get('langs', 'en')
    target_langs = [l.strip() for l in langs_str.split(',') if l.strip()]

    import uuid, shutil
    # 备份原文件
    tmp_id = str(uuid.uuid4())[:8]
    in_path = os.path.join('/tmp', f'xl_in_{tmp_id}.xlsx')
    bak_path = os.path.join('/tmp', f'xl_bak_{tmp_id}.xlsx')
    out_path = os.path.join('/tmp', f'xl_out_{tmp_id}.xlsx')

    f.save(in_path)
    shutil.copy(in_path, bak_path)

    # 读取 Excel
    wb = openpyxl.load_workbook(in_path)
    if in_path.endswith('.xls'):
        # 兼容 xls 格式
        wb.save(in_path.replace('.xls','.xlsx'))
        wb = openpyxl.load_workbook(in_path.replace('.xls','.xlsx'))

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error':'Empty file'}), 400

    header = [str(c).strip() if c else '' for c in rows[0]]
    zh_col = None
    for i, h in enumerate(header):
        if h.lower() in ('zh', '中文', 'chinese'):
            zh_col = i
            break
    if zh_col is None:
        return jsonify({'error':'No zh column found'}), 400

    # 确保目标语言列存在，记录列位置
    lang_cols = {}
    for lang in target_langs:
        if lang == 'zh': continue
        col_idx = None
        for i, h in enumerate(header):
            if h.lower() == lang.lower():
                col_idx = i
                break
        if col_idx is None:
            # 新增列
            col_idx = len(header)
            header.append(lang)
            ws.cell(row=1, column=col_idx+1, value=lang)
        lang_cols[lang] = col_idx

    changes = []
    total_filled = 0
    for r_idx, row in enumerate(rows[1:], start=2):
        zh_val = str(row[zh_col]).strip() if zh_col < len(row) and row[zh_col] else ''
        if not zh_val:
            continue

        translations = translate_multi(zh_val, target_langs)
        for lang in target_langs:
            if lang == 'zh': continue
            col = lang_cols[lang]
            existing = str(row[col]).strip() if col < len(row) and row[col] else ''
            if existing and existing != 'None' and existing != '':
                continue  # 不覆盖已有内容

            result_obj = translations.get(lang, {})
            trans = result_obj.get('value', '') if isinstance(result_obj, dict) else str(result_obj)
            if trans and trans != zh_val:
                ws.cell(row=r_idx, column=col+1, value=trans)
                total_filled += 1
                if total_filled <= 50:
                    changes.append({'row':r_idx, 'zh':zh_val, 'lang':lang, 'trans':trans})

    wb.save(out_path)

    return jsonify({
        'ok': True,
        'filled': total_filled,
        'details': changes,
        'download_id': tmp_id,
        'cols_filled': list(target_langs),
    })

@app.route('/api/download-excel/<dl_id>')
def download_excel(dl_id):
    out_path = os.path.join('/tmp', f'xl_out_{dl_id}.xlsx')
    bak_path = os.path.join('/tmp', f'xl_bak_{dl_id}.xlsx')

    type_param = request.args.get('type', 'translated')
    path = out_path if type_param == 'translated' else bak_path

    if not os.path.exists(path):
        return jsonify({'error':'File expired'}), 404

    from flask import send_file
    fname = 'translated.xlsx' if type_param == 'translated' else 'original_backup.xlsx'
    return send_file(path, as_attachment=True, download_name=fname)

@app.route('/api/dict-stats')
def dict_stats():
    sample = list(dict_zh_to_all.items())[:3]
    return jsonify({
        'total_rows':dict_total_rows,
        'unique_zh':len(dict_zh_to_all),
        'lang_cols':list(dict_lang_cols.keys()),
        'sample':[{'zh':z, 'trans':t} for z,t in sample],
    })

@app.route('/api/langs')
def langs():
    return jsonify(LANG_NAMES)

if __name__ == '__main__':
    print('🚀 i18n Translator Server v2')
    print('   http://localhost:8080')
    app.run(host='0.0.0.0', port=8080, debug=True)
